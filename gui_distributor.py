import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional


class DistributorApp:
    def __init__(self, master):
        self.master = master
        master.title("Распределение остатков")
        master.geometry("650x200")

        self.df = None
        self.result_df = None
        self.saved_file = None  # путь к сохраненному файлу

        # ---------------- Панель загрузки ----------------
        self.label = tk.Label(master, text="Выберите Excel-файл:", font=("Arial", 12))
        self.label.pack(pady=10)

        self.file_label = tk.Label(master, text="Файл не выбран", fg="gray")
        self.file_label.pack(pady=5)

        btn_frame = tk.Frame(master)
        btn_frame.pack(pady=10)

        self.select_button = ttk.Button(btn_frame, text="Выбрать файл", command=self.load_file)
        self.select_button.grid(row=0, column=0, padx=5)

        self.process_button = ttk.Button(btn_frame, text="Распределить", command=self.process_file, state=tk.DISABLED)
        self.process_button.grid(row=0, column=1, padx=5)

        self.save_button = ttk.Button(btn_frame, text="Сохранить результат", command=self.save_result, state=tk.DISABLED)
        self.save_button.grid(row=0, column=2, padx=5)

    # ---------------- Загрузка файла ----------------
    def load_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        
        try:
             raw_df = pd.read_excel(file_path, header=None)

             header_row_index = None
             for i, row in raw_df.iterrows():
                if str(row[0]).strip() == "Магазин":
                    header_row_index = int(i)
                    break

             if header_row_index is None:
                raise ValueError("Не найдена строка с заголовками (Магазин)")

             df = pd.read_excel(file_path, header=header_row_index)
             df.dropna(axis=1, how='all', inplace=True)

             self.df = df
             self.file_label.config(text=file_path, fg="green")
             messagebox.showinfo("Успех", "Файл успешно загружен.")
             self.process_button.config(state=tk.NORMAL)

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл:\n{e}")

    # ---------------- Обработка файла ----------------
    def process_file(self):
        try:
            self.result_df = self.распределить_остатки(self.df)
            self.save_button.config(state=tk.NORMAL)
            messagebox.showinfo("Готово", "Файл обработан. Теперь сохраните результат.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось обработать файл:\n{e}")

    # ---------------- Сохранение ----------------
    def save_result(self):
        try:
            if self.result_df is None:
                messagebox.showwarning("Внимание", "Сначала выполните распределение.")
                return

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if save_path:
                self.result_df.to_excel(save_path, index=False)
                self._format_saved_excel(save_path)  # применяем форматирование
                self.saved_file = save_path
                messagebox.showinfo("Готово", f"Результат сохранён:\n{save_path}")

                # открыть сразу после сохранения
                try:
                    os.startfile(save_path)  # Windows
                except Exception as e:
                    messagebox.showwarning("Внимание", f"Не удалось автоматически открыть файл:\n{e}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{e}")

    # ---------------- Форматирование Excel ----------------
    def _format_saved_excel(self, path: str) -> None:
        wb = load_workbook(path)
        ws: Optional[Worksheet] = wb.active # wb.active может быть None

        if ws is None:
            raise ValueError("Не найден активныйлист в книке Excel")

        header_row = 1
        max_col = ws.max_column
        max_row = ws.max_row

        # увеличенная высота строки заголовков
        ws.row_dimensions[header_row].height = 52

        # стили
        bold_font = Font(bold=True)
        wrap_center = Alignment(wrap_text=True, horizontal="center", vertical="center")
        bottom_line = Side(border_style="medium", color="000000")
        sep_line = Side(border_style="medium", color="000000")

        # форматирование заголовков
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            val = "" if cell.value is None else str(cell.value)
            cell.value = val.replace(" ", "\n")  # перенос строки
            current = cell.border
            cell.border = Border(
                left=current.left,
                right=current.right,
                top=current.top,
                bottom=bottom_line  # нижняя граница по всей строке
            )
            cell.font = bold_font
            cell.alignment = wrap_center

        # вертикальные разделители
        for col_idx in range(1, max_col + 1):
            header_text = str(ws.cell(row=header_row, column=col_idx).value or "")

            # 1. после "… конц"
            if "конц" in header_text:
                for r in range(1, max_row + 1):
                    c = ws.cell(row=r, column=col_idx)
                    cur = c.border
                    c.border = Border(
                        left=cur.left,
                        right=sep_line,
                        top=cur.top,
                        bottom=cur.bottom
                    )

            # 2. после "Нач. остаток на ЦС"
            if header_text.startswith("Нач.\nостаток\nна\nЦС"):
                for r in range(1, max_row + 1):
                    c = ws.cell(row=r, column=col_idx)
                    cur = c.border
                    c.border = Border(
                        left=cur.left,
                        right=sep_line,
                        top=cur.top,
                        bottom=cur.bottom
                    )

        # закрепляем заголовки
        ws.freeze_panes = "A2"

        # включаем автофильтры
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=max_col).coordinate}"

        wb.save(path)

    # ---------------- Логика распределения ----------------
    def распределить_остатки(self, df):
        required_columns = ['Номенклатура', 'Характеристика', 'Магазин', 'Остаток на складе']
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            raise ValueError(f"В файле отсутствуют столбцы: {', '.join(missing)}")

        склады = [
            "Гранд Парк",  "Азия парк Астана",
            "Шымкент «Love is mama»", "Aport East", "Aport West", "ГЦРЧ"
        ]

        grouped = df.groupby(['Номенклатура', 'Характеристика'])
        result_rows = []

        for (номенклатура, характеристика), group in grouped:
            row = {
                "Категория": "Без категории",
                "Номенклатура": номенклатура,
                "Характеристика": характеристика,
                "Нач. остаток на ЦС": 0,
                "Конечный остаток на ЦС": 0
            }

            for склад in склады:
                row[f"{склад} нач"] = 0
                row[f"{склад} кол-во"] = 0
                row[f"{склад} конц"] = 0

            for _, r in group.iterrows():
                магазин = r['Магазин']
                остаток = r['Остаток на складе']

                if магазин == "Центральный склад":
                    row["Нач. остаток на ЦС"] += остаток
                elif магазин in склады:
                    row[f"{магазин} нач"] += остаток

            row["Конечный остаток на ЦС"] = row["Нач. остаток на ЦС"]

            for итерация in range(1, 4):
                for склад in склады:
                    if row[f"{склад} конц"] < итерация and row["Конечный остаток на ЦС"] > 0:
                        row[f"{склад} кол-во"] += 1
                        row[f"{склад} конц"] += 1
                        row["Конечный остаток на ЦС"] -= 1

            result_rows.append(row)

        ordered_columns = ["Категория", "Номенклатура", "Характеристика", "Нач. остаток на ЦС"]
        for склад in склады:
            # ordered_columns += [f"{склад} нач", f"{склад} кол-во", f"{склад} конц"]
            ordered_columns += [f"{склад} кол-во"]
        ordered_columns += ["Конечный остаток на ЦС"]

        return pd.DataFrame(result_rows)[ordered_columns]


## if __name__ == "__main__":
    ##root = tk.Tk()
    ##app = DistributorApp(root)
    ##root.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    app = DistributorApp(root)

    # 👉 здесь вставляем установку иконки
    try:
        root.iconbitmap("favicon.ico")  
    except Exception as e:
        print(f"Не удалось установить иконку: {e}")

    root.mainloop()
